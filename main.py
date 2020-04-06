import os
from openpyxl import load_workbook
from mailmerge import MailMerge


def create_machine_directories():
    """
    The information from each machine shall be stored in its own directory.
    This function handles directory creation for every machine.
    """
    project_root = os.getcwd()
    excel_data = load_workbook(filename=os.path.join(project_root, 'DATA', 'EquipmentList(设备清单).xlsx'))
    excel_sheet = excel_data.active

    for i in range(3, excel_sheet.max_row):
        mach_name = str(excel_sheet.cell(row=i + 1, column=0 + 1).value)
        mach_code = str(excel_sheet.cell(row=i + 1, column=1 + 1).value)
        path = os.path.join(project_root, 'EQUIPMENT', '. '.join((mach_code, mach_name)))
        try:
            os.mkdir(path)
        except OSError:
            print("Failed directory creation at {}".format(path))
        else:
            print("Successfull directory creation at {}".format(path))


def generate_machine_guides():
    """
    Every machine shall have a guide indicating how to operate it.
    This function feeds on an excel file and automatically fills word templates for every machine.
    """
    project_root = os.getcwd()
    template_path = os.path.join(project_root, 'TEMPLATES', "MaintenanceGuide.docx")
    destin_word_loc = os.path.join(project_root, 'EQUIPMENT')
    excel_data = load_workbook(filename=os.path.join(project_root, 'DATA', 'MaintenanceGuideDB.xlsx'))
    excel_sheet = excel_data.active

    # Specifying guides to be created, dealing with duplicates
    guide_list = []
    for xls_row in range(2, excel_sheet.max_row + 1):
        potential_guide = []
        for col in (1, 2, 3, 8):
            potential_guide.append(excel_sheet.cell(row=xls_row, column=col).value)
        if potential_guide not in guide_list:
            guide_list.append(tuple(potential_guide))

    # Fetching data and generating guides
    for guide in guide_list:
        task_counter = 1
        basic_info_created, multiple_machines = False
        tasks_dict = {}
        var_subset = ""

        for xls_row in range(2, excel_sheet.max_row + 1):
            if guide == [excel_sheet.cell(row=xls_row, column=1).value,
                         excel_sheet.cell(row=xls_row, column=2).value,
                         excel_sheet.cell(row=xls_row, column=3).value,
                         excel_sheet.cell(row=xls_row, column=8).value]:

                if not basic_info_created:
                    var_machine = excel_sheet.cell(row=xls_row, column=1).value
                    var_code = excel_sheet.cell(row=xls_row, column=2).value

                    if type(var_code) is float:
                        str(var_code)
                        if "." in var_code:
                            var_code = var_code.replace(".", ",")
                    mach_codes = var_code.split(",")

                    if len(mach_codes) > 1:
                        multiple_machines = True
                    var_freq = excel_sheet.cell(row=xls_row, column=3).value

                    if excel_sheet.cell(row=xls_row, column=8).value:
                        var_subset = excel_sheet.cell(row=xls_row, column=8).value
                        tasks_dict['SUBSET'] = var_subset

                var_resp = excel_sheet.cell(row=xls_row, column=9).value
                var_subset = excel_sheet.cell(row=xls_row, column=8).value
                basic_info_created = True

                var_task_num = excel_sheet.cell(row=xls_row, column=7).value
                var_task_concept = excel_sheet.cell(row=xls_row, column=4).value
                var_task_description = excel_sheet.cell(row=xls_row, column=5).value
                tasks_dict['T' + str(task_counter)] = var_task_num
                tasks_dict['CT' + str(task_counter)] = var_task_concept
                tasks_dict['DT' + str(task_counter)] = var_task_description
                task_counter += 1

                if multiple_machines:
                    for code in mach_codes:
                        with MailMerge(template_path) as document:
                            document.merge(MACHNAME=var_machine, FREQ=var_freq, RESP=var_resp,
                                           CODE=code, **tasks_dict)
                            if var_subset == None:
                                document.write(
                                    destin_word_loc + code + ". " + var_machine + "\\" +
                                    '6_2_508_Guide_' + var_machine + "_" + code + "_" + var_freq + '.docx')
                            else:
                                document.write(
                                    destin_word_loc + code + ". " + var_machine + "\\" +
                                    '6_2_508_Guide_' + var_machine + "_" + str(var_subset) + "_" +
                                    code + "_" + var_freq + '.docx')
                else:
                    with MailMerge(template_path) as document:
                        document.merge(MACHNAME=var_machine, FREQ=var_freq, RESP=var_resp,
                                       CODE=var_code, **tasks_dict)
                        if var_subset is None:
                            document.write(
                                destin_word_loc + var_code + ". " + var_machine + "\\" +
                                '6_2_508_Guide_' + var_machine + "_" + var_code + "_" + var_freq + '.docx')
                        else:
                            document.write(
                                destin_word_loc + var_code + ". " + var_machine + "\\" +
                                '6_2_508_Guide_' + var_machine + "_" + str(var_subset) + "_" +
                                var_code + "_" + var_freq + '.docx')


def generate_machine_datasheets():
    """
    For identification purposes, every machine needs an easily visible sheet gathering all its basic information.
    This function generates that sheet based on information from an excel db.
    """
    project_root = os.getcwd()
    template_path = os.path.join(project_root, 'TEMPLATES', "MachineSheet(机床单).docx")
    destin_word_loc = os.path.join(project_root, 'EQUIPMENT')
    excel_data = load_workbook(filename=os.path.join(project_root, 'DATA', 'EquipmentList(设备清单).xlsx'))
    excel_sheet = excel_data.active

    xls_machine_list = []
    for rownum in range(3, excel_sheet.max_row):
        mach_name = str(excel_sheet.cell(row=rownum + 1, column=0 + 1).value)
        mach_code = str(excel_sheet.cell(row=rownum + 1, column=1 + 1).value)
        folder_name = ". ".join((mach_code, mach_name))
        xls_machine_list.append([folder_name, rownum])

    dir_mach_list = os.listdir(destin_word_loc)
    # print(dir_mach_list)
    for dirname in dir_mach_list:
        curr_mach = dirname
        for i in xls_machine_list:
            if dirname == i[0]:
                mach_row = int(i[1]) + 1
                print("Machine: " + i[0] + " found in db!")
                var_machine = str(excel_sheet.cell(row=mach_row, column=1).value)
                var_code = str(excel_sheet.cell(row=mach_row, column=2).value)
                var_brand = str(excel_sheet.cell(row=mach_row, column=3).value)
                var_model = str(excel_sheet.cell(row=mach_row, column=4).value)
                var_serial = str(excel_sheet.cell(row=mach_row, column=5).value)
                var_weight = str(excel_sheet.cell(row=mach_row, column=6).value)
                var_size = str(excel_sheet.cell(row=mach_row, column=7).value)
                var_purchasedate = str(excel_sheet.cell(row=mach_row, column=8).value)
                var_maintenance = str(excel_sheet.cell(row=mach_row, column=9).value)
                var_power = str(excel_sheet.cell(row=mach_row, column=10).value)
                var_telephone = str(excel_sheet.cell(row=mach_row, column=11).value)
                var_email = str(excel_sheet.cell(row=mach_row, column=12).value)
                var_hasdoc = str(excel_sheet.cell(row=mach_row, column=13).value)
                var_consump = str(excel_sheet.cell(row=mach_row, column=14).value)
                var_location = str(excel_sheet.cell(row=mach_row, column=15).value)
                var_cost = str(excel_sheet.cell(row=mach_row, column=16).value)
                var_comen = str(excel_sheet.cell(row=mach_row, column=17).value)

        with MailMerge(template_path) as document:
            # print(document.get_merge_fields())
            document.merge(CODE=var_code, EMAIL=var_email, MODEL=var_model, COME=var_comen, MAINTENANCE=var_maintenance,
                           LOCATION=var_location, TELEPHONE=var_telephone, CONSUMP=var_consump, MACHINE=var_machine,
                           PURCHASEDATE=var_purchasedate, WEIGHT=var_weight, SERIAL=var_serial, BRAND=var_brand,
                           COST=var_cost, POWER=var_power, SIZE=var_size, HASDOC=var_hasdoc)
            # document.write(dest_loc + '\\6_2_507_Machine Sheet(机床单)_' + curr_mach + '.docx')
            document.write(destin_word_loc + curr_mach + '\\6_2_507_Machine Sheet(机床单)_' + curr_mach + '.docx')


def generate_machine_fault_registers():
    """
    Every marchine shall have a log in which faults are registered.
    This function handles its creation
    """
    project_root = os.getcwd()
    monthly_model_word_loc = os.path.join(project_root, 'TEMPLATES', "MaintenanceRecordMonthly.docx")
    weekly_model_word_loc = os.path.join(project_root, 'TEMPLATES', "MaintenanceRecordWeekly.docx")
    com_model_word_loc = os.path.join(project_root, 'TEMPLATES', "MaintenanceRecordComments.docx")
    destin_word_loc = os.path.join(project_root, 'EQUIPMENT')
    excel_data = load_workbook(filename=os.path.join(project_root, 'DATA', 'MaintenanceGuideDB.xlsx'))
    excel_sheet = excel_data.active

    # SPECIFYING NUMBER OF REGISTERS TO BE CREATED
    register_list = []
    for xls_row in range(2, excel_sheet.max_row + 1):

        potential_register = [excel_sheet.cell(row=xls_row, column=1).value,
                              excel_sheet.cell(row=xls_row, column=2).value,
                              excel_sheet.cell(row=xls_row, column=3).value,
                              excel_sheet.cell(row=xls_row, column=8).value]
        if potential_register not in register_list:
            register_list.append(potential_register)

    # print(register_list)

    # GETTING DATA & CREATING EVERY REGISTER

    for register in register_list:
        task_counter = 1
        basic_info_created = False
        tasks_dict = {}
        var_subset = ""
        multiple_machines = False

        for xls_row in range(2, excel_sheet.max_row + 1):
            if register == [excel_sheet.cell(row=xls_row, column=1).value,
                            excel_sheet.cell(row=xls_row, column=2).value,
                            excel_sheet.cell(row=xls_row, column=3).value,
                            excel_sheet.cell(row=xls_row, column=8).value]:

                if not basic_info_created:
                    var_machine = excel_sheet.cell(row=xls_row, column=1).value
                    var_code = excel_sheet.cell(row=xls_row, column=2).value
                    if type(var_code) is float:
                        str(var_code)
                        if "." in var_code:
                            var_code = var_code.replace(".", ",")
                    mach_codes = var_code.split(",")
                    # print(mach_codes)
                    if len(mach_codes) > 1:
                        multiple_machines = True
                    var_freq = excel_sheet.cell(row=xls_row, column=3).value

                    if not excel_sheet.cell(row=xls_row, column=8).value is None:
                        var_subset = excel_sheet.cell(row=xls_row, column=8).value
                        tasks_dict['SUBSET'] = var_subset

                    var_resp = excel_sheet.cell(row=xls_row, column=9).value
                    var_subset = excel_sheet.cell(row=xls_row, column=8).value
                    basic_info_created = True

                var_task_num = excel_sheet.cell(row=xls_row, column=7).value

                tasks_dict['T' + str(task_counter)] = var_task_num
                task_counter += 1

                # PASSING ALL THE INFORMATION TO THE NEW DOCUMENTS TO BE CREATED
                # MONTHLY REGISTER DOCUMENT CREATION

                if var_freq == "MONTHLY" or var_freq == "EVERY 2 MONTHS":
                    if multiple_machines is True:
                        for code in mach_codes:
                            doc_create(var_machine=var_machine, var_freq=var_freq, var_resp=var_resp,
                                       code=code, var_subset=var_subset, tasks_dict=tasks_dict,
                                       destin_word_loc=destin_word_loc, model_word_loc=monthly_model_word_loc)
                    else:
                        doc_create(var_machine=var_machine, var_freq=var_freq, var_resp=var_resp,
                                   code=var_code, var_subset=var_subset, tasks_dict=tasks_dict,
                                   destin_word_loc=destin_word_loc, model_word_loc=monthly_model_word_loc)
                # WEEKLY REGISTER DOCUMENT CREATION
                elif var_freq == "WEEKLY":
                    if multiple_machines is True:
                        for code in mach_codes:
                            doc_create(var_machine=var_machine, var_freq=var_freq, var_resp=var_resp,
                                       code=code, var_subset=var_subset, tasks_dict=tasks_dict,
                                       destin_word_loc=destin_word_loc, model_word_loc=weekly_model_word_loc)
                    else:
                        doc_create(var_machine=var_machine, var_freq=var_freq, var_resp=var_resp,
                                   code=var_code, var_subset=var_subset, tasks_dict=tasks_dict,
                                   destin_word_loc=destin_word_loc, model_word_loc=weekly_model_word_loc)

                # COMMENTS REGISTER DOCUMENT CREATION

                if multiple_machines is True:
                    for code in mach_codes:
                        doc_create(var_machine=var_machine, var_resp=var_resp,
                                   code=code, comment='Comment_',
                                   destin_word_loc=destin_word_loc, model_word_loc=com_model_word_loc)
                else:
                    doc_create(var_machine=var_machine, var_resp=var_resp,
                               code=var_code, comment='Comment_',
                               destin_word_loc=destin_word_loc, model_word_loc=com_model_word_loc)


def doc_create(var_machine, code, destin_word_loc, model_word_loc, var_freq='', var_resp='', var_subset='',
               tasks_dict={}, comment=''):
    with MailMerge(model_word_loc) as document:
        document.merge(MACHNAME=var_machine, FREQ=var_freq, RESP=var_resp,
                       CODE=code, **tasks_dict)
        if var_subset is None:
            document.write(
                destin_word_loc + code + ". " + var_machine + "\\" +
                '6_2_509_' + comment + 'Register_' + var_machine + "_" + code + "_" + var_freq + '.docx')
        else:
            document.write(
                destin_word_loc + code + ". " + var_machine + "\\" +
                '6_2_509_' + comment + 'Register_' + var_machine + "_" + str(var_subset) + "_" +
                code + "_" + var_freq + '.docx')
        return


if __name__ == "__main__":
    create_machine_directories()